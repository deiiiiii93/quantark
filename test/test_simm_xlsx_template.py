"""Tests for the SIMM xlsx input template (pricing/stress results only)."""
import openpyxl
import pytest

from quantark.util.exceptions import ValidationError
from quantark.simm.config import SIMMConfig
from quantark.simm.taxonomy import ProductClass, RiskClass
from quantark.simm.sensitivity import (
    EquityDeltaSensitivity,
    IRDeltaSensitivity,
    SensitivityCollection,
)
from quantark.simm.engines.aggregation import SIMMCalculator
from quantark.simm.template import (
    PRICING_SHEET,
    STRESS_SHEET,
    calculate_simm_from_xlsx,
    create_simm_input_template,
    load_simm_input,
)


def _fill(path, pricing_rows, stress_rows):
    """Append data rows to a freshly created template."""
    wb = openpyxl.load_workbook(path)
    for row in pricing_rows:
        wb[PRICING_SHEET].append(row)
    for row in stress_rows:
        wb[STRESS_SHEET].append(row)
    wb.save(path)


@pytest.fixture
def template_path(tmp_path):
    path = tmp_path / "simm_input.xlsx"
    create_simm_input_template(path)
    return path


class TestTemplateCreation:
    def test_sheets_and_headers(self, template_path):
        wb = openpyxl.load_workbook(template_path)
        assert "README" in wb.sheetnames
        assert PRICING_SHEET in wb.sheetnames
        assert STRESS_SHEET in wb.sheetnames
        pricing_header = [c.value for c in wb[PRICING_SHEET][1]]
        assert pricing_header == ["TradeId", "ProductClass", "BasePV"]
        stress_header = [c.value for c in wb[STRESS_SHEET][1]]
        assert stress_header == [
            "TradeId", "RiskType", "Qualifier", "Bucket", "Label1", "Label2",
            "ShiftedPV", "ImpliedVol",
        ]


class TestLoading:
    def test_delta_sensitivities_from_pv_differences(self, template_path):
        # Base PV 1,000,000; +1bp USD OIS 5y -> 1,010,000 (PV01 = 10,000);
        # +1% AAPL -> 1,002,000 on the equity trade (s = 2,000).
        _fill(
            template_path,
            pricing_rows=[
                ("SWAP1", "RatesFX", 1_000_000.0),
                ("OPT1", "Equity", 500_000.0),
            ],
            stress_rows=[
                ("SWAP1", "Risk_IRCurve", "USD", "", "5y", "OIS", 1_010_000.0, None),
                ("OPT1", "Risk_Equity", "AAPL", "5", "", "", 502_000.0, None),
            ],
        )
        parsed = load_simm_input(template_path)
        assert len(parsed.sensitivities) == 2

        ir = [s for s in parsed.sensitivities if isinstance(s, IRDeltaSensitivity)][0]
        assert ir.amount == pytest.approx(10_000.0)
        assert ir.vertex == "5y"
        assert ir.effective_product_class == ProductClass.RATES_FX

        eq = [s for s in parsed.sensitivities if isinstance(s, EquityDeltaSensitivity)][0]
        assert eq.amount == pytest.approx(2_000.0)
        assert eq.bucket == 5

    def test_product_class_follows_trade(self, template_path):
        # Paragraph 6: the IR shock of an Equity-product trade belongs to
        # the Equity product class.
        _fill(
            template_path,
            pricing_rows=[("AUTOCALL1", "Equity", 1_000_000.0)],
            stress_rows=[
                ("AUTOCALL1", "Risk_IRCurve", "USD", "", "5y", "OIS", 1_001_000.0, None),
            ],
        )
        parsed = load_simm_input(template_path)
        sens = parsed.sensitivities.sensitivities[0]
        assert sens.risk_class == RiskClass.INTEREST_RATE
        assert sens.effective_product_class == ProductClass.EQUITY

    def test_quoted_vol_required_for_ir_vega(self, template_path):
        _fill(
            template_path,
            pricing_rows=[("SWPT1", "RatesFX", 0.0)],
            stress_rows=[
                ("SWPT1", "Risk_IRVol", "USD", "", "5y", "", 1_000.0, None),
            ],
        )
        with pytest.raises(ValidationError, match="ImpliedVol"):
            load_simm_input(template_path)

    def test_ir_vega_vol_weighting(self, template_path):
        # Raw vega = 1,000 per vol unit; sigma = 0.8 -> amount = 800.
        _fill(
            template_path,
            pricing_rows=[("SWPT1", "RatesFX", 0.0)],
            stress_rows=[
                ("SWPT1", "Risk_IRVol", "USD", "", "5y", "", 1_000.0, 0.8),
            ],
        )
        parsed = load_simm_input(template_path)
        assert parsed.sensitivities.sensitivities[0].amount == pytest.approx(800.0)

    def test_unknown_trade_id_raises(self, template_path):
        _fill(
            template_path,
            pricing_rows=[],
            stress_rows=[
                ("GHOST", "Risk_Equity", "AAPL", "5", "", "", 1_000.0, None),
            ],
        )
        with pytest.raises(ValidationError, match="GHOST"):
            load_simm_input(template_path)


class TestEndToEnd:
    def test_simm_matches_direct_calculation(self, template_path):
        _fill(
            template_path,
            pricing_rows=[
                ("SWAP1", "RatesFX", 1_000_000.0),
                ("OPT1", "Equity", 500_000.0),
            ],
            stress_rows=[
                ("SWAP1", "Risk_IRCurve", "USD", "", "5y", "OIS", 1_010_000.0, None),
                ("SWAP1", "Risk_IRCurve", "USD", "", "10y", "OIS", 1_020_000.0, None),
                ("OPT1", "Risk_Equity", "AAPL", "5", "", "", 502_000.0, None),
            ],
        )
        _, result = calculate_simm_from_xlsx(template_path)

        direct = SIMMCalculator(SIMMConfig()).calculate(SensitivityCollection([
            IRDeltaSensitivity("SWAP1", 10_000.0, currency="USD", tenor=5.0),
            IRDeltaSensitivity("SWAP1", 20_000.0, currency="USD", tenor=10.0),
            EquityDeltaSensitivity("OPT1", 2_000.0, issuer="AAPL", bucket_number=5),
        ]))
        assert result.total_margin == pytest.approx(direct.total_margin)
        assert result.by_product_class[ProductClass.RATES_FX] == pytest.approx(
            direct.by_product_class[ProductClass.RATES_FX])
        assert result.by_product_class[ProductClass.EQUITY] == pytest.approx(26 * 2_000.0)
