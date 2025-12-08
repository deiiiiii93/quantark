"""
Tests for SIMM report generators.
"""
import os
import tempfile
from datetime import date
from unittest.mock import Mock, patch

import pytest

from simm.taxonomy import ProductClass, RiskClass
from simm.results.simm_result import SIMMResult, RiskClassMargin, BucketDetail
from simm.results.attribution import SIMMAttribution, PositionAttribution
from simm.report.html_generator import SIMMHTMLReportGenerator
from simm.report.excel_generator import SIMMExcelReportGenerator
from simm.report.crif_export import (
    export_sensitivities_to_crif,
    write_crif_csv,
    create_crif_template,
)


class TestSIMMHTMLReportGenerator:
    """Test SIMMHTMLReportGenerator."""

    def test_init(self):
        """Test initializing the generator."""
        generator = SIMMHTMLReportGenerator(include_charts=True)
        assert generator.include_charts is True

    def test_generate_basic_report(self):
        """Test generating a basic HTML report."""
        # Create a simple result
        result = SIMMResult(
            total_simm=1000.0,
            calculation_currency="USD",
            calculation_date=date(2024, 1, 15),
            simm_version="2.6",
            product_class_simm={
                ProductClass.RATES_FX: 500.0,
                ProductClass.EQUITY: 300.0,
                ProductClass.CREDIT: 150.0,
                ProductClass.COMMODITY: 50.0,
            },
            risk_class_margin={},
            addon_amount=0.0,
        )

        # Create generator
        generator = SIMMHTMLReportGenerator(include_charts=False)

        # Generate report to temp file
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
            output_path = f.name

        try:
            result_path = generator.generate(result, output_path, include_charts=False)

            assert os.path.exists(result_path)

            # Check file content
            with open(result_path, 'r') as f:
                content = f.read()

            assert 'SIMM Calculation Report' in content
            assert 'Total SIMM' in content
            assert '1,000.00' in content or '1000.00' in content

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_generate_report_with_charts(self):
        """Test generating HTML report with charts."""
        result = SIMMResult(
            total_simm=1000.0,
            calculation_currency="USD",
            calculation_date=date(2024, 1, 15),
            simm_version="2.6",
            product_class_simm={pc: 250.0 for pc in ProductClass},
            risk_class_margin={},
            addon_amount=0.0,
        )

        generator = SIMMHTMLReportGenerator(include_charts=True)

        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
            output_path = f.name

        try:
            result_path = generator.generate(result, output_path, include_charts=True)

            assert os.path.exists(result_path)

            with open(result_path, 'r') as f:
                content = f.read()

            assert 'SIMM Calculation Report' in content

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_get_currency_symbol(self):
        """Test currency symbol extraction."""
        generator = SIMMHTMLReportGenerator()

        assert generator._get_currency_symbol('USD') == '$'
        assert generator._get_currency_symbol('EUR') == '€'
        assert generator._get_currency_symbol('GBP') == '£'
        assert generator._get_currency_symbol('JPY') == '¥'
        assert generator._get_currency_symbol('XXX') == 'XXX '

    def test_css_styles(self):
        """Test CSS styles are generated."""
        generator = SIMMHTMLReportGenerator()
        css = generator._get_css_styles()

        assert '.container' in css
        '.report-header' in css
        assert '.section' in css
        assert '.data-table' in css

    @patch('simm.report.html_generator.PLOTLY_AVAILABLE', False)
    def test_generate_without_plotly(self):
        """Test generating report without Plotly."""
        result = SIMMResult(
            total_simm=1000.0,
            calculation_currency="USD",
            calculation_date=date(2024, 1, 15),
            simm_version="2.6",
            product_class_simm={pc: 250.0 for pc in ProductClass},
            risk_class_margin={},
            addon_amount=0.0,
        )

        generator = SIMMHTMLReportGenerator(include_charts=True)

        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
            output_path = f.name

        try:
            result_path = generator.generate(result, output_path, include_charts=True)

            assert os.path.exists(result_path)

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


class TestSIMMExcelReportGenerator:
    """Test SIMMExcelReportGenerator."""

    @pytest.mark.skipif(not pytest.importorskip("openpyxl").__name__, reason="openpyxl not installed")
    def test_init(self):
        """Test initializing the generator."""
        generator = SIMMExcelReportGenerator()
        assert generator is not None

    @pytest.mark.skipif(not pytest.importorskip("openpyxl").__name__, reason="openpyxl not installed")
    def test_generate_basic_report(self):
        """Test generating a basic Excel report."""
        # Create a simple result
        result = SIMMResult(
            total_simm=1000.0,
            calculation_currency="USD",
            calculation_date=date(2024, 1, 15),
            simm_version="2.6",
            product_class_simm={
                ProductClass.RATES_FX: 500.0,
                ProductClass.EQUITY: 300.0,
                ProductClass.CREDIT: 150.0,
                ProductClass.COMMODITY: 50.0,
            },
            risk_class_margin={},
            addon_amount=0.0,
        )

        # Create generator
        generator = SIMMExcelReportGenerator()

        # Generate report to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        try:
            result_path = generator.generate(result, output_path)

            assert os.path.exists(result_path)

            # Verify it's a valid Excel file by trying to read it
            import openpyxl
            wb = openpyxl.load_workbook(result_path)
            assert "Summary" in wb.sheetnames

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    @pytest.mark.skipif(not pytest.importorskip("openpyxl").__name__, reason="openpyxl not installed")
    def test_generate_to_bytes(self):
        """Test generating report to bytes."""
        result = SIMMResult(
            total_simm=1000.0,
            calculation_currency="USD",
            calculation_date=date(2024, 1, 15),
            simm_version="2.6",
            product_class_simm={pc: 250.0 for pc in ProductClass},
            risk_class_margin={},
            addon_amount=0.0,
        )

        generator = SIMMExcelReportGenerator()
        excel_bytes = generator.generate_to_bytes(result)

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # Verify it's a valid Excel file
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        assert "Summary" in wb.sheetnames


class TestCRIFExport:
    """Test CRIF export functionality."""

    def test_create_crif_template(self):
        """Test creating a CRIF template."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
            output_path = f.name

        try:
            result_path = create_crif_template(output_path)

            assert os.path.exists(result_path)

            # Verify file content
            with open(result_path, 'r') as f:
                content = f.read()

            assert 'Valuation_Date' in content
            assert 'Trade_ID' in content
            assert 'Risk_Type' in content

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_write_crif_csv(self):
        """Test writing CRIF to CSV."""
        from simm.crif import CRIFRecord
        from simm.taxonomy import SensitivityType

        # Create mock CRIF records
        records = [
            CRIFRecord(
                trade_id="TRADE_001",
                valuation_date=date(2024, 1, 15),
                risk_type=SensitivityType.RISK_IR_CURVE,
                qualifier="USD",
                bucket="USD",
                label1="",
                label2="",
                amount=1000000.0,
                amount_currency="USD",
            ),
        ]

        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
            output_path = f.name

        try:
            write_crif_csv(records, output_path, date(2024, 1, 15))

            assert os.path.exists(output_path)

            # Verify file content
            with open(output_path, 'r') as f:
                content = f.read()

            assert 'TRADE_001' in content
            assert 'Risk_IRCurve' in content

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_sensitivities_to_crif_file(self):
        """Test exporting sensitivities to CRIF file."""
        from simm.sensitivity import SensitivityCollection

        # Create mock sensitivities
        sensitivities = SensitivityCollection([])

        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = f.name

        try:
            result_path = export_sensitivities_to_crif(
                sensitivities,
                trade_id="TRADE_001",
                valuation_date=date(2024, 1, 15),
                output_path=output_path,
            )

            assert result_path == output_path
            assert os.path.exists(output_path)

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)


if __name__ == "__main__":
    pytest.main([__file__])
