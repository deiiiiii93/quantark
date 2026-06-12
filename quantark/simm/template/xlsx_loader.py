"""
SIMM input template loader.

Reads a workbook created by ``create_simm_input_template`` (and filled
with pricing/stress results) and converts each StressResults row into a
SIMM sensitivity using the shift conventions of ISDA SIMM v2.6
Sections C.2 and C.3:

    s = ShiftedPV - BasePV

For vega rows the difference is the raw vega dV/dsigma for a 1-unit vol
shift; it is converted to the vol-weighted amount sigma_kj * dV/dsigma:

- Equity / Commodity / FX: sigma_kj derived from the delta risk weight
  (paragraph 10(b)), applied automatically.
- Interest Rate / Credit: sigma_kj is the quoted ATM implied volatility
  (paragraph 10(a)) and must be supplied in the ImpliedVol column.
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl

from quantark.util.exceptions import ValidationError
from quantark.simm.config import SIMMConfig
from quantark.simm.taxonomy import (
    CREDIT_TENORS,
    CREDIT_TENOR_LABELS,
    IR_TENORS,
    IR_TENOR_LABELS,
    IRSubCurve,
    MarginType,
    ProductClass,
    RiskClass,
    SensitivityType,
)
from quantark.simm.sensitivity import (
    AnySensitivity,
    BaseCorrSensitivity,
    CommodityDeltaSensitivity,
    CommodityVegaSensitivity,
    CreditDeltaSensitivity,
    CreditVegaSensitivity,
    EquityDeltaSensitivity,
    EquityVegaSensitivity,
    FXDeltaSensitivity,
    FXVegaSensitivity,
    IRDeltaSensitivity,
    IRInflationDeltaSensitivity,
    IRVegaSensitivity,
    IRXCcyBasisSensitivity,
    SensitivityCollection,
    vol_weighted_vega_commodity,
    vol_weighted_vega_equity,
    vol_weighted_vega_fx,
)
from quantark.simm.template.xlsx_template import (
    PRICING_COLUMNS,
    PRICING_SHEET,
    STRESS_COLUMNS,
    STRESS_SHEET,
)

# Risk types whose vega requires a quoted implied volatility
# (paragraph 10(a)).
_QUOTED_VOL_TYPES = {
    SensitivityType.RISK_IR_VOL,
    SensitivityType.RISK_INFLATION_VOL,
    SensitivityType.RISK_CREDIT_VOL,
    SensitivityType.RISK_CREDIT_NQ_VOL,
}


@dataclass
class TradePricing:
    """Base pricing result for one trade."""
    trade_id: str
    product_class: ProductClass
    base_pv: float


@dataclass
class SIMMTemplateInput:
    """Parsed content of a SIMM input template workbook.

    Attributes:
        trades: Base pricing results per trade id.
        sensitivities: Sensitivities derived from the stress results.
        warnings: Non-fatal issues encountered while parsing.
    """
    trades: Dict[str, TradePricing] = field(default_factory=dict)
    sensitivities: SensitivityCollection = field(default_factory=SensitivityCollection)
    warnings: List[str] = field(default_factory=list)


def _parse_tenor_label(label: str, risk_class: RiskClass) -> float:
    """Convert a vertex label to years."""
    label = str(label).strip().lower()
    if risk_class in (RiskClass.CREDIT_QUALIFYING, RiskClass.CREDIT_NON_QUALIFYING):
        if label in CREDIT_TENOR_LABELS:
            return CREDIT_TENORS[CREDIT_TENOR_LABELS.index(label)]
    if label in IR_TENOR_LABELS:
        return IR_TENORS[IR_TENOR_LABELS.index(label)]
    raise ValidationError(f"Unknown tenor label: {label!r}")


def _parse_bucket(raw) -> Union[int, str]:
    if raw is None or str(raw).strip() == "":
        return 1
    text = str(raw).strip()
    if text.lower() in ("residual", "res", "-1"):
        return -1
    return int(float(text))


def _parse_sub_curve(raw) -> IRSubCurve:
    if raw is None or str(raw).strip() == "":
        return IRSubCurve.OIS
    text = str(raw).strip().upper()
    for sc in IRSubCurve:
        if sc.value.upper() == text:
            return sc
    raise ValidationError(f"Unknown IR sub-curve: {raw!r}")


def _rows(sheet, columns) -> List[Dict[str, object]]:
    header = [cell.value for cell in sheet[1]]
    missing = [c for c in columns if c not in header]
    if missing:
        raise ValidationError(
            f"Sheet {sheet.title!r} is missing columns: {missing}"
        )
    index = {name: header.index(name) for name in columns}
    out = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        out.append({name: row[i] for name, i in index.items()})
    return out


def _convert_row(
    row: Dict[str, object],
    trade: TradePricing,
    row_num: int,
) -> AnySensitivity:
    """Convert one StressResults row into a sensitivity."""
    risk_type_raw = str(row["RiskType"]).strip()
    sensitivity_type = next(
        (st for st in SensitivityType if st.value == risk_type_raw), None
    )
    if sensitivity_type is None:
        raise ValidationError(f"Row {row_num}: unknown RiskType {risk_type_raw!r}")

    qualifier = str(row["Qualifier"] or "").strip()
    if not qualifier:
        raise ValidationError(f"Row {row_num}: Qualifier is required")

    shifted_pv = row["ShiftedPV"]
    if shifted_pv is None:
        raise ValidationError(f"Row {row_num}: ShiftedPV is required")
    s = float(shifted_pv) - trade.base_pv  # Sections C.2 / C.3

    risk_class = sensitivity_type.risk_class
    margin_type = sensitivity_type.margin_type

    # Vol-weight vega amounts (paragraph 10(a)-(b)).
    if margin_type == MarginType.VEGA:
        if sensitivity_type in _QUOTED_VOL_TYPES:
            implied_vol = row.get("ImpliedVol")
            if implied_vol is None or str(implied_vol).strip() == "":
                raise ValidationError(
                    f"Row {row_num}: ImpliedVol is required for {risk_type_raw} "
                    "(paragraph 10(a))"
                )
            s = float(implied_vol) * s

    label1 = row.get("Label1")
    common: Dict[str, Any] = dict(
        trade_id=trade.trade_id,
        amount=s,
        product_class=trade.product_class,
    )

    if sensitivity_type == SensitivityType.RISK_IR_CURVE:
        return IRDeltaSensitivity(
            currency=qualifier,
            tenor=_parse_tenor_label(str(label1), risk_class),
            sub_curve=_parse_sub_curve(row.get("Label2")),
            **common,
        )
    elif sensitivity_type == SensitivityType.RISK_INFLATION:
        return IRInflationDeltaSensitivity(currency=qualifier, **common)
    elif sensitivity_type == SensitivityType.RISK_XCCY_BASIS:
        return IRXCcyBasisSensitivity(currency=qualifier, **common)
    elif sensitivity_type in (SensitivityType.RISK_IR_VOL, SensitivityType.RISK_INFLATION_VOL):
        return IRVegaSensitivity(
            currency=qualifier,
            option_tenor=_parse_tenor_label(str(label1), risk_class),
            is_inflation=sensitivity_type == SensitivityType.RISK_INFLATION_VOL,
            **common,
        )
    elif sensitivity_type in (SensitivityType.RISK_CREDIT_Q, SensitivityType.RISK_CREDIT_NQ):
        return CreditDeltaSensitivity(
            issuer=qualifier,
            bucket_number=_parse_bucket(row.get("Bucket")),
            tenor=_parse_tenor_label(str(label1), risk_class),
            is_qualifying=sensitivity_type == SensitivityType.RISK_CREDIT_Q,
            **common,
        )
    elif sensitivity_type in (SensitivityType.RISK_CREDIT_VOL, SensitivityType.RISK_CREDIT_NQ_VOL):
        return CreditVegaSensitivity(
            issuer=qualifier,
            bucket_number=_parse_bucket(row.get("Bucket")),
            option_tenor=_parse_tenor_label(str(label1), risk_class),
            is_qualifying=sensitivity_type == SensitivityType.RISK_CREDIT_VOL,
            **common,
        )
    elif sensitivity_type == SensitivityType.RISK_BASE_CORR:
        return BaseCorrSensitivity(index_name=qualifier, **common)
    elif sensitivity_type == SensitivityType.RISK_EQUITY:
        return EquityDeltaSensitivity(
            issuer=qualifier,
            bucket_number=_parse_bucket(row.get("Bucket")),
            **common,
        )
    elif sensitivity_type == SensitivityType.RISK_EQUITY_VOL:
        bucket = _parse_bucket(row.get("Bucket"))
        common["amount"] = vol_weighted_vega_equity(s, bucket)
        return EquityVegaSensitivity(
            issuer=qualifier,
            bucket_number=bucket,
            option_tenor=_parse_tenor_label(str(label1), risk_class) if label1 else 1.0,
            **common,
        )
    elif sensitivity_type == SensitivityType.RISK_COMMODITY:
        return CommodityDeltaSensitivity(
            commodity_name=qualifier,
            bucket_number=int(_parse_bucket(row.get("Bucket"))),
            **common,
        )
    elif sensitivity_type == SensitivityType.RISK_COMMODITY_VOL:
        bucket = int(_parse_bucket(row.get("Bucket")))
        common["amount"] = vol_weighted_vega_commodity(s, bucket)
        return CommodityVegaSensitivity(
            commodity_name=qualifier,
            bucket_number=bucket,
            option_tenor=_parse_tenor_label(str(label1), risk_class) if label1 else 1.0,
            **common,
        )
    elif sensitivity_type == SensitivityType.RISK_FX:
        return FXDeltaSensitivity(currency=qualifier, **common)
    elif sensitivity_type == SensitivityType.RISK_FX_VOL:
        common["amount"] = vol_weighted_vega_fx(s, qualifier)
        return FXVegaSensitivity(
            currency_pair=qualifier,
            option_tenor=_parse_tenor_label(str(label1), risk_class) if label1 else 1.0,
            **common,
        )

    raise ValidationError(f"Row {row_num}: unsupported RiskType {risk_type_raw!r}")


def load_simm_input(path: Union[str, Path]) -> SIMMTemplateInput:
    """Load a filled SIMM input template workbook.

    Args:
        path: Path to the filled .xlsx template.

    Returns:
        SIMMTemplateInput with base pricing results and the derived
        sensitivities.

    Raises:
        ValidationError: On malformed sheets or rows.
    """
    wb = openpyxl.load_workbook(Path(path), data_only=True)
    for sheet_name in (PRICING_SHEET, STRESS_SHEET):
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Workbook is missing sheet {sheet_name!r}")

    result = SIMMTemplateInput()

    for i, row in enumerate(_rows(wb[PRICING_SHEET], PRICING_COLUMNS), start=2):
        trade_id = str(row["TradeId"] or "").strip()
        if not trade_id:
            raise ValidationError(f"PricingResults row {i}: TradeId is required")
        pc_raw = str(row["ProductClass"] or "").strip()
        product_class = next(
            (pc for pc in ProductClass if pc.value.lower() == pc_raw.lower()), None
        )
        if product_class is None:
            raise ValidationError(
                f"PricingResults row {i}: unknown ProductClass {pc_raw!r}"
            )
        base_pv = row["BasePV"]
        if base_pv is None:
            raise ValidationError(f"PricingResults row {i}: BasePV is required")
        result.trades[trade_id] = TradePricing(
            trade_id=trade_id,
            product_class=product_class,
            base_pv=float(base_pv),
        )

    for i, row in enumerate(_rows(wb[STRESS_SHEET], STRESS_COLUMNS), start=2):
        trade_id = str(row["TradeId"] or "").strip()
        trade = result.trades.get(trade_id)
        if trade is None:
            raise ValidationError(
                f"StressResults row {i}: TradeId {trade_id!r} has no "
                f"PricingResults entry"
            )
        result.sensitivities.add(_convert_row(row, trade, i))

    return result


def calculate_simm_from_xlsx(
    path: Union[str, Path],
    config: Optional[SIMMConfig] = None,
) -> Tuple["SIMMTemplateInput", "object"]:
    """Load a filled template and calculate SIMM.

    Args:
        path: Path to the filled .xlsx template.
        config: Optional SIMM configuration.

    Returns:
        Tuple of (parsed template input, SIMMAggregationResult).
    """
    from quantark.simm.engines.aggregation import SIMMCalculator

    template_input = load_simm_input(path)
    calculator = SIMMCalculator(config)
    result = calculator.calculate(template_input.sensitivities)
    return template_input, result
