"""
SIMM input template workbook generator.

Creates an Excel workbook whose only inputs are pricing and stress
results: a base present value per trade and one shifted present value
per risk-factor shock. This is exactly the sensitivity definition of
ISDA SIMM v2.6 Sections C.2 and C.3:

- Interest Rate / Credit delta: shift the rate/spread up 1bp
- Equity / Commodity / FX delta: shift the price/rate up 1% (relative)
- Vega: shift the implied at-the-money volatility up 1 unit
- Base Correlation: shift the base correlation up 1 percentage point

The loader (quantark.simm.template.xlsx_loader) computes
s = ShiftedPV - BasePV for each row and runs the SIMM calculation.
"""

from typing import Union
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from quantark.simm.taxonomy import (
    CREDIT_TENOR_LABELS,
    IR_TENOR_LABELS,
    IRSubCurve,
    ProductClass,
    SensitivityType,
)

# Sheet names
README_SHEET = "README"
PRICING_SHEET = "PricingResults"
STRESS_SHEET = "StressResults"

# Column layouts
PRICING_COLUMNS = ("TradeId", "ProductClass", "BasePV")
STRESS_COLUMNS = (
    "TradeId", "RiskType", "Qualifier", "Bucket", "Label1", "Label2",
    "ShiftedPV", "ImpliedVol",
)

# Shift convention per CRIF risk type (documented on the README sheet and
# enforced by the loader).
SHIFT_CONVENTIONS = (
    ("Risk_IRCurve", "Sub-curve (Label2) shifted up 1bp at the tenor (Label1); Qualifier = currency"),
    ("Risk_Inflation", "Flat inflation rate shifted up 1bp; Qualifier = currency"),
    ("Risk_XCcyBasis", "Cross-currency basis spread shifted up 1bp; Qualifier = currency"),
    ("Risk_IRVol", "ATM swaption vol shifted up 1 unit at expiry (Label1); requires ImpliedVol; Qualifier = currency"),
    ("Risk_InflationVol", "ATM inflation swaption vol shifted up 1 unit at expiry (Label1); requires ImpliedVol; Qualifier = currency"),
    ("Risk_CreditQ", "Credit spread shifted up 1bp at the tenor (Label1); Qualifier = issuer/seniority; Bucket 1-12 or Residual"),
    ("Risk_CreditVol", "ATM credit vol shifted up 1 unit at expiry (Label1); requires ImpliedVol"),
    ("Risk_CreditNonQ", "Credit spread shifted up 1bp at the tenor (Label1); Bucket 1-2 or Residual"),
    ("Risk_CreditVolNonQ", "ATM credit vol shifted up 1 unit at expiry (Label1); requires ImpliedVol"),
    ("Risk_BaseCorr", "Base correlation shifted up 1 percentage point; Qualifier = index family"),
    ("Risk_Equity", "Equity price shifted up 1% (relative); Qualifier = equity name; Bucket 1-12 or Residual"),
    ("Risk_EquityVol", "Implied vol shifted up 1 vol point at expiry (Label1); sigma_kj derived from risk weights"),
    ("Risk_Commodity", "Commodity price shifted up 1% (relative); Qualifier = commodity name; Bucket 1-17"),
    ("Risk_CommodityVol", "Implied vol shifted up 1 vol point at expiry (Label1); sigma_kj derived from risk weights"),
    ("Risk_FX", "FX rate vs calculation currency shifted up 1% (relative); Qualifier = currency"),
    ("Risk_FXVol", "Implied vol shifted up 1 vol point at expiry (Label1); Qualifier = currency pair (e.g. EURUSD)"),
)

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)

# Number of rows covered by the data validations.
_VALIDATION_ROWS = 5000


def _style_header(sheet, columns) -> None:
    for col, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(col)].width = max(14, len(name) + 4)
    sheet.freeze_panes = "A2"


def _list_validation(sheet, values, column_letter) -> None:
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(values) + '"',
        allow_blank=True,
        showErrorMessage=True,
    )
    sheet.add_data_validation(dv)
    dv.add(f"{column_letter}2:{column_letter}{_VALIDATION_ROWS}")


def create_simm_input_template(path: Union[str, Path]) -> Path:
    """Create the SIMM pricing/stress-result input template workbook.

    The workbook contains:
    - README: shift conventions and usage instructions
    - PricingResults: TradeId | ProductClass | BasePV
    - StressResults: one row per risk-factor shock with the shifted PV

    Args:
        path: Output .xlsx path.

    Returns:
        The path the template was written to.
    """
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # README
    # ------------------------------------------------------------------
    readme = wb.active
    assert readme is not None
    readme.title = README_SHEET
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 120

    readme["A1"] = "ISDA SIMM v2.6 input template (pricing / stress results only)"
    readme["A1"].font = _TITLE_FONT

    rows = [
        ("", ""),
        ("How to use", "1. Fill PricingResults with one row per trade: TradeId, ProductClass and the base present value (BasePV) in the calculation currency."),
        ("", "2. Fill StressResults with one row per risk-factor shock: re-price the trade with the single shift described below and record the shifted PV."),
        ("", "3. The SIMM sensitivity is computed as s = ShiftedPV - BasePV; no Greeks are required."),
        ("", "4. Load with quantark.simm.template.load_simm_input(path) or calculate_simm_from_xlsx(path)."),
        ("", ""),
        ("Conventions", "All PVs in the SIMM calculation currency (default USD). ProductClass follows the trade (paragraph 6): e.g. the IR shock of an equity derivative is reported with ProductClass = Equity."),
        ("", "Label1 = tenor / option expiry vertex. IR vertices: " + ", ".join(IR_TENOR_LABELS) + ". Credit vertices: " + ", ".join(CREDIT_TENOR_LABELS) + "."),
        ("", "Label2 = IR sub-curve (" + ", ".join(sc.value for sc in IRSubCurve) + "); leave blank otherwise."),
        ("", "ImpliedVol = the at-the-money implied volatility sigma_kj, required only for Risk_IRVol / Risk_InflationVol / Risk_CreditVol / Risk_CreditVolNonQ rows (paragraph 10(a))."),
        ("", ""),
        ("Shift per RiskType", ""),
    ]
    r = 2
    for label, text in rows:
        readme.cell(row=r, column=1, value=label).font = Font(bold=True)
        readme.cell(row=r, column=2, value=text)
        r += 1
    for risk_type, desc in SHIFT_CONVENTIONS:
        readme.cell(row=r, column=1, value=risk_type).font = Font(bold=True)
        readme.cell(row=r, column=2, value=desc)
        r += 1

    # ------------------------------------------------------------------
    # PricingResults
    # ------------------------------------------------------------------
    pricing = wb.create_sheet(PRICING_SHEET)
    _style_header(pricing, PRICING_COLUMNS)
    _list_validation(pricing, [pc.value for pc in ProductClass], "B")

    # ------------------------------------------------------------------
    # StressResults
    # ------------------------------------------------------------------
    stress = wb.create_sheet(STRESS_SHEET)
    _style_header(stress, STRESS_COLUMNS)
    _list_validation(stress, [st.value for st in SensitivityType], "B")
    _list_validation(stress, list(IR_TENOR_LABELS), "E")
    _list_validation(stress, [sc.value for sc in IRSubCurve], "F")

    out = Path(path)
    wb.save(out)
    return out
