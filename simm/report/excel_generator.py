"""
SIMM Excel Report Generator.

This module generates Excel workbooks for SIMM results with multiple sheets
and professional formatting.
"""
from typing import Dict, List, Any, Optional
from datetime import date
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..results.simm_result import SIMMResult
from ..results.attribution import SIMMAttribution
from ..taxonomy import ProductClass, RiskClass


class SIMMExcelReportGenerator:
    """Generate Excel reports for SIMM results."""

    def __init__(self):
        """Initialize report generator."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel report generation")

    def generate(
        self,
        result: SIMMResult,
        output_path: str,
    ) -> str:
        """Generate Excel workbook.

        Args:
            result: SIMMResult to report on.
            output_path: Output Excel file path.

        Returns:
            Path to generated Excel file.
        """
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, result)
        self._create_product_class_sheet(wb, result)
        self._create_risk_class_sheets(wb, result)
        self._create_position_attribution_sheet(wb, result)
        self._create_crif_sheet(wb, result)

        # Save workbook
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb: openpyxl.Workbook, result: SIMMResult):
        """Create summary sheet with key metrics.

        Args:
            wb: Excel workbook.
            result: SIMMResult to report on.
        """
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = 'SIMM Calculation Summary'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')

        # Summary data
        row = 3
        summary_data = [
            ('Calculation Date', result.calculation_date),
            ('Currency', result.calculation_currency),
            ('SIMM Version', result.simm_version),
            ('Total SIMM', result.total_simm),
            ('Add-ons', result.addon_amount),
            ('Execution Time (s)', result.execution_time_seconds),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].number_format = '#,##0.00' if isinstance(value, (int, float)) else 'General'
            row += 1

        # Product class summary
        row += 2
        ws[f'A{row}'] = 'Product Class Breakdown'
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Headers
        headers = ['Product Class', 'SIMM Amount', '% of Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        row += 1

        # Data
        for pc, amount in sorted(result.product_class_simm.items(), key=lambda x: x[1], reverse=True):
            pct = (amount / result.total_simm * 100) if result.total_simm != 0 else 0
            ws[f'A{row}'] = pc.value
            ws[f'B{row}'] = amount
            ws[f'C{row}'] = pct / 100
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'C{row}'].number_format = '0.00%'
            row += 1

        # Auto-fit columns
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_product_class_sheet(self, wb: openpyxl.Workbook, result: SIMMResult):
        """Create product class breakdown sheet.

        Args:
            wb: Excel workbook.
            result: SIMMResult to report on.
        """
        ws = wb.create_sheet("Product Class Breakdown")

        # Title
        ws['A1'] = 'Product Class Detailed Breakdown'
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        for pc in ProductClass:
            if pc not in result.risk_class_margin:
                continue

            ws[f'A{row}'] = pc.value
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1

            # Headers
            headers = ['Risk Class', 'Delta Margin', 'Vega Margin', 'Curvature Margin', 'Base Corr', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

            row += 1

            # Data
            for rc, rc_margin in result.risk_class_margin[pc].items():
                ws[f'A{row}'] = rc.value
                ws[f'B{row}'] = rc_margin.delta_margin
                ws[f'C{row}'] = rc_margin.vega_margin
                ws[f'D{row}'] = rc_margin.curvature_margin
                ws[f'E{row}'] = rc_margin.base_corr_margin
                ws[f'F{row}'] = rc_margin.total_margin

                for col in range(2, 7):
                    ws.cell(row=row, column=col).number_format = '#,##0.00'

                row += 1

            row += 2  # Space between product classes

        # Auto-fit columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_risk_class_sheets(self, wb: openpyxl.Workbook, result: SIMMResult):
        """Create per-risk-class detail sheets.

        Args:
            wb: Excel workbook.
            result: SIMMResult to report on.
        """
        for rc in RiskClass:
            sheet_name = f"{rc.value} Detail"
            ws = wb.create_sheet(sheet_name)

            # Title
            ws['A1'] = f'{rc.value} Risk Class Detail'
            ws['A1'].font = Font(size=14, bold=True)

            row = 3

            for pc in ProductClass:
                if pc not in result.risk_class_margin:
                    continue

                if rc not in result.risk_class_margin[pc]:
                    continue

                rc_margin = result.risk_class_margin[pc][rc]

                ws[f'A{row}'] = f'{pc.value} - {rc.value}'
                ws[f'A{row}'].font = Font(size=12, bold=True)
                row += 1

                # Margin components
                ws[f'A{row}'] = 'Margin Components'
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

                components = [
                    ('Delta Margin', rc_margin.delta_margin),
                    ('Vega Margin', rc_margin.vega_margin),
                    ('Curvature Margin', rc_margin.curvature_margin),
                    ('Base Correlation Margin', rc_margin.base_corr_margin),
                    ('Total Margin', rc_margin.total_margin),
                ]

                for label, value in components:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].number_format = '#,##0.00'
                    row += 1

                # Bucket detail
                if rc_margin.bucket_detail:
                    row += 1
                    ws[f'A{row}'] = 'Bucket Detail'
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1

                    # Headers
                    headers = ['Bucket', 'K Value', 'WS Sum', 'Concentration Factor', 'S Value']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                    row += 1

                    # Bucket data
                    for bucket, bucket_detail in sorted(
                        rc_margin.bucket_detail.items(),
                        key=lambda x: x[1].k_value,
                        reverse=True
                    ):
                        ws[f'A{row}'] = str(bucket)
                        ws[f'B{row}'] = bucket_detail.k_value
                        ws[f'C{row}'] = bucket_detail.ws_sum
                        ws[f'D{row}'] = bucket_detail.concentration_factor
                        ws[f'E{row}'] = bucket_detail.s_value

                        for col in range(2, 6):
                            ws.cell(row=row, column=col).number_format = '#,##0.00'

                        row += 1

                row += 3  # Space between product classes

        # Auto-fit columns for all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_position_attribution_sheet(self, wb: openpyxl.Workbook, result: SIMMResult):
        """Create position attribution sheet.

        Args:
            wb: Excel workbook.
            result: SIMMResult to report on.
        """
        ws = wb.create_sheet("Position Attribution")

        # Title
        ws['A1'] = 'Position-Level Attribution'
        ws['A1'].font = Font(size=14, bold=True)

        if not isinstance(result.attribution, SIMMAttribution):
            ws['A3'] = 'No attribution data available'
            return

        row = 3

        # Headers
        headers = [
            'Position ID',
            'Trade ID',
            'Underlying',
            'Delta Contribution',
            'Vega Contribution',
            'Curvature Contribution',
            'Total Contribution',
            '% of Total'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        row += 1

        # Data
        sorted_positions = sorted(
            result.attribution.by_position.values(),
            key=lambda p: p.total_contribution,
            reverse=True
        )

        for pos in sorted_positions:
            ws[f'A{row}'] = pos.position_id
            ws[f'B{row}'] = pos.trade_id or ''
            ws[f'C{row}'] = pos.underlying
            ws[f'D{row}'] = pos.delta_contribution
            ws[f'E{row}'] = pos.vega_contribution
            ws[f'F{row}'] = pos.curvature_contribution
            ws[f'G{row}'] = pos.total_contribution
            ws[f'H{row}'] = pos.pct_of_total / 100

            for col in range(4, 8):
                ws.cell(row=row, column=col).number_format = '#,##0.00'

            ws[f'H{row}'].number_format = '0.00%'
            row += 1

        # Auto-fit columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_crif_sheet(self, wb: openpyxl.Workbook, result: SIMMResult):
        """Create CRIF export sheet.

        Args:
            wb: Excel workbook.
            result: SIMMResult to report on.
        """
        ws = wb.create_sheet("CRIF Export")

        # Title
        ws['A1'] = 'CRIF Format Export'
        ws['A1'].font = Font(size=14, bold=True)

        ws['A3'] = 'This sheet contains sensitivities in CRIF format'
        ws['A3'].font = Font(italic=True, color='666666')

        # Note: In a full implementation, you would export the actual sensitivities
        # For now, we'll just add a placeholder

        row = 5

        # Headers (simplified CRIF format)
        headers = [
            'Trade_ID',
            'Risk_Type',
            'Bucket',
            'Sensitivity',
            'Qualifier',
            'Currency',
            'Tenor'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

        row += 1

        ws[f'A{row}'] = 'N/A - Export actual sensitivities from result'
        ws[f'A{row}'].font = Font(italic=True, color='666666')

        # Auto-fit columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def generate_to_bytes(self, result: SIMMResult) -> bytes:
        """Generate Excel report to bytes.

        Args:
            result: SIMMResult to report on.

        Returns:
            Excel file as bytes.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, result)
        self._create_product_class_sheet(wb, result)
        self._create_risk_class_sheets(wb, result)
        self._create_position_attribution_sheet(wb, result)
        self._create_crif_sheet(wb, result)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()
